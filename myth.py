import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="Al Madina | Performance Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
#MainMenu, footer, header { visibility: hidden; }
.stDeployButton { display: none; }
[data-testid="stSidebar"] { background: #0f1117; border-right: 1px solid #1e2130; }
[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stMultiSelect label,
[data-testid="stSidebar"] .stFileUploader label { color: #94a3b8 !important; font-size: 11px !important; text-transform: uppercase; letter-spacing: .08em; }
[data-testid="stSidebar"] hr { border-color: #1e2130; }
.main .block-container { padding: 0 2rem 2rem; max-width: 1400px; }
.top-bar { background: linear-gradient(135deg,#0f1117 0%,#1a1f2e 100%); border-bottom:1px solid #1e2130; padding:16px 28px; margin:-1rem -2rem 2rem; display:flex; align-items:center; justify-content:space-between; }
.top-bar-brand { display:flex; align-items:center; gap:12px; }
.logo-circle { width:38px;height:38px;border-radius:10px;background:linear-gradient(135deg,#f97316,#dc2626);display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;color:white; }
.top-bar-brand h1 { font-size:17px;font-weight:600;color:#f1f5f9;margin:0; }
.top-bar-brand span { font-size:11px;color:#64748b; }
.top-bar-meta { font-size:11px;color:#475569; }
.section-label { font-size:10px;font-weight:600;letter-spacing:.12em;text-transform:uppercase;color:#94a3b8;margin:1.5rem 0 .6rem;padding-bottom:.5rem;border-bottom:1px solid #e2e8f0; }
.kpi-card { background:white;border:1px solid #e2e8f0;border-radius:12px;padding:14px 18px;position:relative;overflow:hidden; }
.kpi-card::before { content:'';position:absolute;top:0;left:0;right:0;height:3px; }
.kpi-card.orange::before{background:linear-gradient(90deg,#f97316,#fb923c);}
.kpi-card.green::before{background:linear-gradient(90deg,#10b981,#34d399);}
.kpi-card.red::before{background:linear-gradient(90deg,#ef4444,#f87171);}
.kpi-card.blue::before{background:linear-gradient(90deg,#3b82f6,#60a5fa);}
.kpi-card.purple::before{background:linear-gradient(90deg,#8b5cf6,#a78bfa);}
.kpi-card.amber::before{background:linear-gradient(90deg,#f59e0b,#fbbf24);}
.kpi-label{font-size:10px;color:#94a3b8;font-weight:600;text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px;}
.kpi-value{font-size:24px;font-weight:600;color:#0f172a;line-height:1;}
.kpi-sub{font-size:11px;color:#64748b;margin-top:3px;}
.outlet-card{background:white;border:1px solid #e2e8f0;border-radius:14px;padding:18px;margin-bottom:10px;}
.outlet-header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px;}
.outlet-name{font-size:14px;font-weight:600;color:#0f172a;}
.outlet-area{font-size:11px;color:#94a3b8;margin-top:2px;}
.outlet-badge{font-size:11px;font-weight:600;padding:3px 10px;border-radius:20px;white-space:nowrap;display:inline-block;}
.badge-good{background:#d1fae5;color:#065f46;}
.badge-warn{background:#fef3c7;color:#92400e;}
.badge-bad{background:#fee2e2;color:#991b1b;}
.prog-wrap{margin:3px 0 7px;}
.prog-label{display:flex;justify-content:space-between;font-size:11px;color:#64748b;margin-bottom:3px;}
.prog-track{height:5px;background:#f1f5f9;border-radius:3px;overflow:hidden;}
.prog-fill{height:100%;border-radius:3px;}
.compare-row{display:flex;gap:6px;flex-wrap:wrap;margin:6px 0;}
.compare-chip{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:5px 10px;font-size:11px;color:#475569;display:flex;gap:5px;align-items:center;}
.chip-val{font-weight:600;color:#0f172a;}
.alert-box{border-radius:10px;padding:11px 14px;margin:7px 0;font-size:13px;display:flex;gap:10px;align-items:flex-start;}
.alert-red{background:#fff1f2;border:1px solid #fecdd3;color:#9f1239;}
.alert-amber{background:#fffbeb;border:1px solid #fed7aa;color:#92400e;}
.alert-green{background:#f0fdf4;border:1px solid #bbf7d0;color:#14532d;}
.alert-blue{background:#eff6ff;border:1px solid #bfdbfe;color:#1e3a8a;}
.alert-icon{font-size:15px;flex-shrink:0;}
.stTabs [data-baseweb="tab-list"]{gap:4px;border-bottom:2px solid #e2e8f0;}
.stTabs [data-baseweb="tab"]{background:transparent;border:none;border-radius:8px 8px 0 0;padding:8px 16px;font-size:13px;font-weight:500;color:#64748b;}
.stTabs [aria-selected="true"]{background:white;color:#f97316 !important;border-bottom:2px solid #f97316;}
.stDownloadButton>button{background:linear-gradient(135deg,#f97316,#dc2626) !important;color:white !important;border:none !important;border-radius:10px !important;font-weight:600 !important;width:100% !important;}
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
NUMERIC_COLS = [
    "Subtotal","Packaging charges","Minimum order value fee","Vendor Refunds",
    "plugins.reports.order_details_report.customer_fee_total",
    "Tax Charge","Online Payment Fee","Discount Funded by you","Voucher Funded by you",
    "Commission","Operational Charges","Ads Fee","Marketing Fees",
    "Avoidable cancellation fee","Estimated earnings",
    "Cash amount already collected by you","Amount owed back to Talabat",
    "Payout Amount","Talabat-Funded Discount","Talabat-Funded Voucher",
    "Total Discount","Total Voucher","Tax Amount"
]
DATE_COLS = [
    "Order received at","Accepted at","Estimated ready to pick up time",
    "Ready to pick up at","Rider near pickup at","In delivery at",
    "Estimated delivery time","Delivered at","Cancelled at"
]
OUTLET_COLORS = {
    "Liwan": "#f97316",
    "Dubai Investment Park": "#3b82f6",
    "Oud Metha": "#8b5cf6",
    "Naif": "#ef4444",
    "Al Muteena": "#10b981",
    "Al Hamriya": "#f59e0b",
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_area(name):
    n = str(name).lower()
    if "liwan" in n: return "Liwan"
    if "dip" in n or "investment park" in n: return "Dubai Investment Park"
    if "oud metha" in n: return "Oud Metha"
    if "naif" in n: return "Naif"
    if "muteena" in n: return "Al Muteena"
    if "hamriya" in n: return "Al Hamriya"
    return "Other"

def safe_div(a, b, pct=False):
    if b == 0: return 0
    return (a/b*100) if pct else a/b

def outlet_metrics(df_sub):
    d = df_sub[df_sub["Order status"] == "Delivered"]
    c = df_sub[df_sub["Order status"] == "Cancelled"]
    total = len(df_sub)
    gmv   = d["Subtotal"].sum()
    diff  = (d["Delivered at"] - d["Order received at"]).dt.total_seconds()/60
    diff  = diff[diff > 0]
    prep  = (d["Ready to pick up at"] - d["Accepted at"]).dt.total_seconds()/60
    prep  = prep[prep > 0]
    lm    = (d["Delivered at"] - d["In delivery at"]).dt.total_seconds()/60
    lm    = lm[lm > 0]
    return {
        "total": total, "delivered": len(d), "cancelled": len(c),
        "can_rate": safe_div(len(c), total, pct=True),
        "del_rate": safe_div(len(d), total, pct=True),
        "gmv": gmv, "payout": d["Payout Amount"].sum(),
        "commission": d["Commission"].sum(),
        "op_charges": d["Operational Charges"].sum(),
        "online_fee": d["Online Payment Fee"].sum(),
        "avg_order": safe_div(gmv, len(d)),
        "del_time": round(diff.mean(),1) if len(diff)>0 else None,
        "prep_time": round(prep.mean(),1) if len(prep)>0 else None,
        "last_mile": round(lm.mean(),1) if len(lm)>0 else None,
        "pro_pct": safe_div(len(d[d["Is Pro Order"]=="Y"]), len(d), pct=True),
        "online_pct": safe_div(len(d[d["Payment type"]=="Online"]), len(d), pct=True),
        "complaints": len(d[d["Has Complaint?"]=="Y"]),
        "complaint_rate": safe_div(len(d[d["Has Complaint?"]=="Y"]), len(d), pct=True),
        "lost_gmv": c["Subtotal"].sum(),
    }

def can_badge(rate):
    if rate <= 8:  return "badge-good","Low"
    if rate <= 20: return "badge-warn","Medium"
    return "badge-bad","High"

def del_badge(mins):
    if mins is None: return "badge-warn","N/A"
    if mins <= 35:   return "badge-good","Fast"
    if mins <= 45:   return "badge-warn","Average"
    return "badge-bad","Slow"

def kpi(label, value, sub="", color="orange"):
    st.markdown(f"""
    <div class="kpi-card {color}">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        {"<div class='kpi-sub'>"+sub+"</div>" if sub else ""}
    </div>""", unsafe_allow_html=True)

def progress_bar(label, value, max_val, color="#f97316", suffix=""):
    pct = min(safe_div(value,max_val,pct=True), 100)
    st.markdown(f"""
    <div class="prog-wrap">
        <div class="prog-label"><span>{label}</span><span>{value}{suffix}</span></div>
        <div class="prog-track"><div class="prog-fill" style="width:{pct:.1f}%;background:{color};"></div></div>
    </div>""", unsafe_allow_html=True)

def alert(msg, type_="blue", icon="ℹ"):
    st.markdown(f"""
    <div class="alert-box alert-{type_}">
        <span class="alert-icon">{icon}</span><span>{msg}</span>
    </div>""", unsafe_allow_html=True)

# ── Data Loader ───────────────────────────────────────────────────────────────
def _col_to_idx(col_str):
    idx = 0
    for c in col_str:
        idx = idx * 26 + (ord(c.upper()) - ord('A') + 1)
    return idx - 1

def _parse_sheet_xml(sheet_bytes):
    """
    Pure-Python xlsx sheet parser.
    Handles inlineStr cells (no sharedStrings.xml needed).
    Returns a pd.DataFrame with no header applied.
    """
    import re as _re
    rows_xml = _re.findall(rb'<row\b[^>]*>(.*?)</row>', sheet_bytes, _re.DOTALL)
    all_rows = {}
    max_col  = 0

    for row_xml in rows_xml:
        cells = _re.findall(rb'<c\b([^>]*?)>(.*?)</c>', row_xml, _re.DOTALL)
        for attrs, content in cells:
            ref_m = _re.search(rb'r="([A-Z]+)(\d+)"', attrs)
            if not ref_m:
                continue
            col_i = _col_to_idx(ref_m.group(1).decode())
            row_i = int(ref_m.group(2)) - 1
            max_col = max(max_col, col_i)

            is_val = _re.search(rb'<is><t[^>]*>(.*?)</t></is>', content, _re.DOTALL)
            v_val  = _re.search(rb'<v>(.*?)</v>', content)

            if is_val:
                val = is_val.group(1).decode('utf-8', 'replace')
            elif v_val:
                raw_v = v_val.group(1).decode('utf-8', 'replace')
                try:
                    val = float(raw_v) if '.' in raw_v else int(raw_v)
                except Exception:
                    val = raw_v
            else:
                val = ''

            all_rows.setdefault(row_i, {})[col_i] = val

    n_rows = max(all_rows.keys()) + 1 if all_rows else 0
    n_cols = max_col + 1
    matrix = [
        [all_rows.get(ri, {}).get(ci, '') for ci in range(n_cols)]
        for ri in range(n_rows)
    ]
    return pd.DataFrame(matrix)


def _fix_and_read_openpyxl(file_bytes):
    """Rewrite xl/styles.xml to fix invalid alignment values, then read with openpyxl."""
    import zipfile, re as _re

    VALID_VERT  = {b"top", b"center", b"bottom", b"justify", b"distributed"}
    VALID_HORIZ = {b"general", b"left", b"center", b"right", b"fill",
                   b"justify", b"centerContinuous", b"distributed"}

    def fix_attr(xml, attr, valid, fallback):
        pat = _re.compile(rb'(' + attr + rb'=")([^"]*?)(")')
        return pat.sub(
            lambda m: m.group(0) if m.group(2) in valid
                      else m.group(1) + fallback + m.group(3),
            xml
        )

    buf_out = BytesIO()
    with zipfile.ZipFile(BytesIO(file_bytes), "r") as zin, \
         zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = fix_attr(data, b"vertical",   VALID_VERT,  b"bottom")
                data = fix_attr(data, b"horizontal",  VALID_HORIZ, b"general")
            zout.writestr(item.filename, data)

    buf_out.seek(0)
    return pd.read_excel(buf_out, engine="openpyxl", header=0)


def _read_excel_robust(file_bytes):
    """
    Multi-strategy Excel reader — ordered so the most reliable runs first.
    Strategy 1: Pure-Python stdlib XML (zero deps, always works for .xlsx)
    Strategy 2: calamine (fastest, immune to stylesheet bugs)
    Strategy 3: openpyxl after repairing corrupt xl/styles.xml
    Strategy 4: openpyxl raw
    Strategy 5: xlrd (legacy .xls)
    """
    import zipfile

    # ── 1. Pure-Python stdlib XML — works on ANY Python, zero dependencies ──
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as z:
            names = [i.filename for i in z.infolist()]
            if "xl/worksheets/sheet1.xml" in names:
                sheet_bytes = z.read("xl/worksheets/sheet1.xml")
                df = _parse_sheet_xml(sheet_bytes)
                if len(df) > 1:   # sanity check: got actual rows
                    return df
    except Exception:
        pass

    # ── 2. calamine ────────────────────────────────────────────────────────
    try:
        import python_calamine  # noqa
        return pd.read_excel(BytesIO(file_bytes), engine="calamine", header=0)
    except Exception:
        pass

    # ── 3. openpyxl after styles.xml repair ────────────────────────────────
    try:
        return _fix_and_read_openpyxl(file_bytes)
    except Exception:
        pass

    # ── 4. openpyxl raw ────────────────────────────────────────────────────
    try:
        return pd.read_excel(BytesIO(file_bytes), engine="openpyxl", header=0)
    except Exception:
        pass

    # ── 5. xlrd (legacy .xls) ──────────────────────────────────────────────
    try:
        return pd.read_excel(BytesIO(file_bytes), engine="xlrd", header=0)
    except Exception as e:
        raise RuntimeError(
            "Could not read the uploaded file. "
            "Please open it in Excel → File → Save As → .xlsx and re-upload. "
            f"Detail: {e}"
        )



def _style_df(df, heat_col=None, heat_col2=None, reverse=False):
    """Apply inline bar-style highlighting without matplotlib."""
    def bar_color(val, col_vals, color):
        try:
            mn, mx = col_vals.min(), col_vals.max()
            if mx == mn:
                pct = 50
            else:
                pct = (val - mn) / (mx - mn) * 100
            if reverse:
                pct = 100 - pct
            return f"background: linear-gradient(90deg, {color} {pct:.0f}%, transparent {pct:.0f}%); color: #0f172a;"
        except Exception:
            return ""

    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    if heat_col and heat_col in df.columns:
        col_vals = pd.to_numeric(df[heat_col], errors="coerce").fillna(0)
        styles[heat_col] = col_vals.apply(lambda v: bar_color(v, col_vals, "rgba(249,115,22,0.25)"))
    if heat_col2 and heat_col2 in df.columns:
        col_vals2 = pd.to_numeric(df[heat_col2], errors="coerce").fillna(0)
        styles[heat_col2] = col_vals2.apply(lambda v: bar_color(v, col_vals2, "rgba(239,68,68,0.25)"))
    return df.style.apply(lambda _: styles, axis=None)


def _style_df_full(df, color="rgba(249,115,22,0.25)"):
    """Highlight every numeric column as a heatmap row."""
    def row_style(row):
        styles = []
        for col in df.columns:
            try:
                col_vals = pd.to_numeric(df[col], errors="coerce").dropna()
                v = pd.to_numeric(row[col], errors="coerce")
                if col_vals.empty or pd.isna(v):
                    styles.append("")
                    continue
                mn, mx = col_vals.min(), col_vals.max()
                pct = 0 if mx == mn else (v - mn) / (mx - mn) * 100
                styles.append(f"background: linear-gradient(90deg, {color} {pct:.0f}%, transparent {pct:.0f}%); color: #0f172a;")
            except Exception:
                styles.append("")
        return styles
    return df.style.apply(row_style, axis=1)



@st.cache_data
def load_data(file_bytes):
    df_raw = _read_excel_robust(file_bytes)

    # Find the real header row — search first 5 rows for "Order status"
    header_row = None
    for i in range(min(5, len(df_raw))):
        row_vals = [str(v) for v in df_raw.iloc[i].values]
        if "Order status" in row_vals:
            header_row = i
            break

    if header_row is None:
        # Last resort: use row 0 as header
        header_row = 0

    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = [str(c) for c in df_raw.iloc[header_row].tolist()]
    df.reset_index(drop=True, inplace=True)

    # Strip any completely empty rows
    df = df.dropna(how="all").reset_index(drop=True)

    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    for col in DATE_COLS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    if "Restaurant name" not in df.columns:
        raise RuntimeError(
            "Uploaded file does not appear to be a Talabat order report. "
            "Expected column 'Restaurant name' not found."
        )

    df["_area"] = df["Restaurant name"].apply(get_area)
    df["_date"] = df["Order received at"].dt.date
    df["_hour"] = df["Order received at"].dt.hour
    df["_dow"]  = df["Order received at"].dt.day_name()
    return df

# ── Excel Builder ─────────────────────────────────────────────────────────────
def build_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    C_DARK="0F1117"; C_OR="F97316"; C_GR="10B981"; C_RE="EF4444"
    C_BL="3B82F6"; C_PU="8B5CF6"; C_AM="F59E0B"
    C_WH="FFFFFF"; C_LG="F8FAFC"; C_MG="E2E8F0"
    OHEX={"Liwan":C_OR,"Dubai Investment Park":C_BL,"Oud Metha":C_PU,
          "Naif":C_RE,"Al Muteena":C_GR,"Al Hamriya":C_AM}

    def fl(h): return PatternFill("solid",start_color=h,fgColor=h)
    def fn(bold=False,size=9,color=C_DARK): return Font(bold=bold,size=size,color=color,name="Calibri")
    def bd():
        s=Side(style="thin",color=C_MG)
        return Border(left=s,right=s,top=s,bottom=s)
    def ct(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def lf(): return Alignment(horizontal="left",vertical="center")
    def rg(): return Alignment(horizontal="right",vertical="center")

    def title_row(ws,row,text,color=C_DARK,nc=14):
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=nc)
        c=ws.cell(row=row,column=1,value=text)
        c.font=Font(bold=True,size=14,color=C_WH,name="Calibri")
        c.fill=fl(color); c.alignment=ct(); ws.row_dimensions[row].height=28
        return row+1

    def sec(ws,row,text,color=C_OR,nc=14):
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=nc)
        c=ws.cell(row=row,column=1,value=text)
        c.font=Font(bold=True,size=10,color=C_WH,name="Calibri")
        c.fill=fl(color); c.alignment=lf(); ws.row_dimensions[row].height=20
        return row+1

    def hdr(ws,row,headers,color=C_OR,widths=None):
        for ci,h in enumerate(headers,1):
            c=ws.cell(row=row,column=ci,value=h)
            c.font=fn(bold=True,size=8,color=C_WH); c.fill=fl(color)
            c.alignment=ct(); c.border=bd()
            if widths and ci-1<len(widths):
                ws.column_dimensions[get_column_letter(ci)].width=widths[ci-1]
        return row+1

    def rows(ws,row,data,nums=None):
        for ri,rd in enumerate(data):
            bg=C_LG if ri%2==0 else C_WH
            for ci,val in enumerate(rd,1):
                c=ws.cell(row=row+ri,column=ci,value=val)
                c.font=fn(size=8); c.fill=fl(bg); c.border=bd()
                c.alignment=rg() if (nums and ci in nums) else lf()
        return row+len(data)+1

    outlets_all = sorted(df["_area"].unique())
    delivered = df[df["Order status"]=="Delivered"]
    cancelled = df[df["Order status"]=="Cancelled"]
    date_min = df["Order received at"].min()
    date_max = df["Order received at"].max()

    outlet_data = []
    for area in outlets_all:
        s = df[df["_area"]==area]
        if len(s)>0:
            outlet_data.append((area, s["Restaurant name"].iloc[0], outlet_metrics(s)))
    outlet_data.sort(key=lambda x: x[2]["gmv"], reverse=True)

    # SHEET 1: Group Summary
    ws = wb.active; ws.title="Group Summary"; ws.sheet_view.showGridLines=False
    row = title_row(ws,1,"AL MADINA HYPERMARKET  ·  GROUP PERFORMANCE REPORT",C_DARK)
    ws.merge_cells("A2:N2")
    s2=ws["A2"]
    s2.value=f"Period: {date_min.strftime('%d %b %Y') if pd.notna(date_min) else '—'} – {date_max.strftime('%d %b %Y') if pd.notna(date_max) else '—'}   |   {len(outlet_data)} Outlets   |   Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    s2.font=Font(size=8,color="94A3B8",name="Calibri",italic=True)
    s2.fill=fl(C_LG); s2.alignment=ct(); ws.row_dimensions[2].height=14
    row=4
    m=outlet_metrics(df)
    row=sec(ws,row,"  GROUP KEY METRICS",C_DARK)
    hdr(ws,row,["Metric","Value","Metric","Value","Metric","Value"],C_DARK,[22,16,22,16,22,16])
    row+=1
    krows=[
        ["Total Orders",m["total"],"Delivered",f"{m['delivered']} ({m['del_rate']:.1f}%)","Cancelled",f"{m['cancelled']} ({m['can_rate']:.1f}%)"],
        ["Gross Revenue (GMV)",f"AED {m['gmv']:,.0f}","Total Payout",f"AED {m['payout']:,.0f}","Commission",f"AED {m['commission']:,.0f}"],
        ["Avg Order Value",f"AED {m['avg_order']:,.1f}","Avg Delivery",f"{m['del_time']} min" if m['del_time'] else "N/A","Complaints",f"{m['complaints']} ({m['complaint_rate']:.1f}%)"],
        ["Online Orders %",f"{m['online_pct']:.0f}%","Pro Orders %",f"{m['pro_pct']:.0f}%","Lost to Cancellations",f"AED {m['lost_gmv']:,.0f}"],
    ]
    for ri,rd in enumerate(krows):
        bg=C_LG if ri%2==0 else C_WH
        for ci,val in enumerate(rd,1):
            c=ws.cell(row=row+ri,column=ci,value=val)
            is_lbl=ci%2==1
            c.font=Font(bold=is_lbl,size=9,color=C_OR if is_lbl else C_DARK,name="Calibri")
            c.fill=fl(bg); c.alignment=lf(); c.border=bd()
    row+=len(krows)+2

    row=sec(ws,row,"  OUTLET LEAGUE TABLE (sorted by GMV)",C_OR)
    hdr_cols=["Rank","Outlet","Area","Orders","Delivered","Cancelled","Cancel %","GMV (AED)","Payout (AED)","Commission","Avg Order","Del Time","Complaints","Pro %"]
    hdr(ws,row,hdr_cols,C_OR,[6,26,22,8,10,10,10,14,14,12,11,11,11,10])
    row+=1
    tbl_rows=[]
    for rank,(area,fname,om) in enumerate(outlet_data,1):
        tbl_rows.append([rank,fname,area,om["total"],om["delivered"],om["cancelled"],
            f"{om['can_rate']:.1f}%",f"{om['gmv']:,.0f}",f"{om['payout']:,.0f}",
            f"{om['commission']:,.0f}",f"{om['avg_order']:,.1f}",
            f"{om['del_time']:.0f}" if om['del_time'] else "N/A",
            om["complaints"],f"{om['pro_pct']:.0f}%"])
    rows(ws,row,tbl_rows,{4,5,6,8,9,10,11,12})
    row+=len(tbl_rows)+2

    row=sec(ws,row,"  DAILY GROUP TREND",C_BL)
    hdr(ws,row,["Date","Total","Delivered","Cancelled","Cancel %","GMV (AED)","Payout (AED)"],C_BL,[14,10,12,12,12,14,14])
    row+=1
    dg=df.groupby("_date").agg(total=("Order ID","count"),
        delivered=("Order status",lambda x:(x=="Delivered").sum()),
        cancelled=("Order status",lambda x:(x=="Cancelled").sum()),
        gmv=("Subtotal","sum")).reset_index()
    dp=delivered.groupby(delivered["_date"])["Payout Amount"].sum().reset_index()
    dp.columns=["_date","payout"]; dg=dg.merge(dp,on="_date",how="left").fillna(0)
    dg["can_pct"]=dg.apply(lambda r:safe_div(r["cancelled"],r["total"],pct=True),axis=1)
    dr=[[str(r["_date"]),int(r["total"]),int(r["delivered"]),int(r["cancelled"]),
         f"{r['can_pct']:.1f}%",f"{r['gmv']:,.0f}",f"{r['payout']:,.0f}"] for _,r in dg.iterrows()]
    rows(ws,row,dr,{2,3,4,6,7})
    ws.freeze_panes="A5"

    # SHEET 2: Outlet Comparison
    ws2=wb.create_sheet("Outlet Comparison"); ws2.sheet_view.showGridLines=False
    row=title_row(ws2,1,"OUTLET COMPARISON MATRIX",C_OR); row+=1
    row=sec(ws2,row,"  METRIC VS METRIC",C_OR)
    ws2.cell(row=row,column=1,value="Metric").font=Font(bold=True,size=8,color=C_WH,name="Calibri")
    ws2.cell(row=row,column=1).fill=fl(C_OR)
    ws2.cell(row=row,column=1).border=bd()
    ws2.column_dimensions["A"].width=26
    for ci,(area,_,_) in enumerate(outlet_data,2):
        c=ws2.cell(row=row,column=ci,value=area)
        c.font=Font(bold=True,size=8,color=C_WH,name="Calibri")
        c.fill=fl(OHEX.get(area,C_OR)); c.alignment=ct(); c.border=bd()
        ws2.column_dimensions[get_column_letter(ci)].width=18
    row+=1
    metrics_list=[
        ("Total Orders","total"),("Delivered","delivered"),("Cancelled","cancelled"),
        ("Cancellation Rate %","can_rate"),("GMV (AED)","gmv"),("Payout (AED)","payout"),
        ("Commission (AED)","commission"),("Op Charges (AED)","op_charges"),
        ("Avg Order Value","avg_order"),("Avg Delivery Time","del_time"),
        ("Avg Prep Time","prep_time"),("Avg Last Mile","last_mile"),
        ("Pro Order %","pro_pct"),("Online Payment %","online_pct"),
        ("Complaints","complaints"),("Complaint Rate %","complaint_rate"),
        ("Lost GMV (Cancels)","lost_gmv"),
    ]
    for ri,(label,key) in enumerate(metrics_list):
        bg=C_LG if ri%2==0 else C_WH
        lc=ws2.cell(row=row+ri,column=1,value=label)
        lc.font=fn(bold=True,size=8); lc.fill=fl(bg); lc.alignment=lf(); lc.border=bd()
        for ci,(_,_,om) in enumerate(outlet_data,2):
            val=om.get(key,0) or 0
            if key in("can_rate","del_rate","pro_pct","online_pct","complaint_rate"): disp=f"{val:.1f}%"
            elif key in("gmv","payout","commission","op_charges","avg_order","lost_gmv"): disp=f"{val:,.0f}"
            elif key in("del_time","prep_time","last_mile"): disp=f"{val:.0f} min" if val else "N/A"
            else: disp=str(int(val)) if isinstance(val,float) else str(val)
            vc=ws2.cell(row=row+ri,column=ci,value=disp)
            vc.font=fn(size=8); vc.fill=fl(bg); vc.alignment=rg(); vc.border=bd()
    row+=len(metrics_list)+2

    row=sec(ws2,row,"  CANCELLATION BREAKDOWN BY OUTLET",C_RE)
    hdr(ws2,row,["Outlet","Cancelled","Vendor","Customer","Rider","Platform","Item N/A","Fraudulent","Other","Lost GMV (AED)"],
        C_RE,[22,10,10,10,10,10,12,12,12,16])
    row+=1
    can_r=[]
    for area,_,_ in outlet_data:
        sc=cancelled[cancelled["_area"]==area]
        owners=sc["Cancellation owner"].value_counts()
        can_r.append([area,len(sc),
            owners.get("Vendor",0),owners.get("Customer",0),owners.get("Rider",0),
            int(sc["Cancellation owner"].str.strip().eq("Platform").sum()),
            int(sc["Cancellation reason"].str.contains("Item not available",na=False).sum()),
            int(sc["Cancellation reason"].str.contains("Fraudulent",na=False).sum()),
            len(sc)-int(sc["Cancellation reason"].str.contains("Item not available|Fraudulent",na=False).sum()),
            f"{sc['Subtotal'].sum():,.0f}"])
    rows(ws2,row,can_r,{2,3,4,5,6,7,8,9,10})

    # SHEETS 3+: Per outlet
    for area,fname,om in outlet_data:
        safe_t=area[:28]; ws_o=wb.create_sheet(safe_t)
        ws_o.sheet_view.showGridLines=False
        color=OHEX.get(area,C_OR)
        row=title_row(ws_o,1,f"OUTLET REPORT  ·  {area.upper()}",color)
        ws_o.merge_cells("A2:N2"); s2=ws_o["A2"]
        s2.value=f"{fname}   |   {date_min.strftime('%d %b') if pd.notna(date_min) else '?'} – {date_max.strftime('%d %b %Y') if pd.notna(date_max) else '?'}"
        s2.font=Font(size=8,color="94A3B8",name="Calibri",italic=True)
        s2.fill=fl(C_LG); s2.alignment=lf(); ws_o.row_dimensions[2].height=14
        row=4
        sub_df=df[df["_area"]==area]; sub_d=sub_df[sub_df["Order status"]=="Delivered"]
        sub_c=sub_df[sub_df["Order status"]=="Cancelled"]
        row=sec(ws_o,row,"  KEY METRICS",color)
        kd=[["Total Orders",om["total"],"Delivered",f"{om['delivered']} ({om['del_rate']:.1f}%)","Cancelled",f"{om['cancelled']} ({om['can_rate']:.1f}%)"],
            ["GMV",f"AED {om['gmv']:,.0f}","Payout",f"AED {om['payout']:,.0f}","Commission",f"AED {om['commission']:,.0f}"],
            ["Avg Order",f"AED {om['avg_order']:,.1f}","Avg Delivery",f"{om['del_time']:.0f} min" if om['del_time'] else "N/A","Avg Prep",f"{om['prep_time']:.0f} min" if om['prep_time'] else "N/A"],
            ["Pro Orders",f"{om['pro_pct']:.0f}%","Online Payments",f"{om['online_pct']:.0f}%","Complaints",f"{om['complaints']} ({om['complaint_rate']:.1f}%)"]]
        for ri,rd in enumerate(kd):
            bg=C_LG if ri%2==0 else C_WH
            for ci,val in enumerate(rd,1):
                c=ws_o.cell(row=row+ri,column=ci,value=val)
                is_lbl=ci%2==1
                c.font=Font(bold=is_lbl,size=9,color=color if is_lbl else C_DARK,name="Calibri")
                c.fill=fl(bg); c.alignment=lf(); c.border=bd()
        row+=len(kd)+2
        row=sec(ws_o,row,"  DAILY PERFORMANCE",color)
        hdr(ws_o,row,["Date","Total","Delivered","Cancelled","Cancel %","GMV (AED)","Payout (AED)"],color,[14,10,12,12,12,14,14])
        row+=1
        ds=sub_df.groupby("_date").agg(total=("Order ID","count"),
            delivered=("Order status",lambda x:(x=="Delivered").sum()),
            cancelled=("Order status",lambda x:(x=="Cancelled").sum()),
            gmv=("Subtotal","sum")).reset_index()
        dp2=sub_d.groupby(sub_d["_date"])["Payout Amount"].sum().reset_index()
        dp2.columns=["_date","payout"]; ds=ds.merge(dp2,on="_date",how="left").fillna(0)
        ds["can_pct"]=ds.apply(lambda r:safe_div(r["cancelled"],r["total"],pct=True),axis=1)
        dr2=[[str(r["_date"]),int(r["total"]),int(r["delivered"]),int(r["cancelled"]),
              f"{r['can_pct']:.1f}%",f"{r['gmv']:,.0f}",f"{r['payout']:,.0f}"] for _,r in ds.iterrows()]
        rows(ws_o,row,dr2,{2,3,4,6,7}); row+=len(dr2)+2
        if len(sub_c)>0:
            row=sec(ws_o,row,"  CANCELLATION DETAIL",C_RE)
            hdr(ws_o,row,["Reason","Count","Owner","Lost GMV (AED)"],C_RE,[36,10,18,16])
            row+=1
            cd=sub_c.groupby(["Cancellation reason","Cancellation owner"]).agg(
                Count=("Order ID","count"),Lost=("Subtotal","sum")).reset_index().sort_values("Count",ascending=False)
            cdr=[[str(r["Cancellation reason"]),int(r["Count"]),str(r["Cancellation owner"]).strip(),f"{r['Lost']:,.0f}"] for _,r in cd.iterrows()]
            rows(ws_o,row,cdr,{2,4}); row+=len(cdr)+2
        row=sec(ws_o,row,"  FINANCIAL BREAKDOWN",color)
        hdr(ws_o,row,["Item","Total (AED)","Per Order","% of GMV"],color,[30,16,14,13])
        row+=1
        fin=[(l,v) for l,v in [("GMV",om["gmv"]),("Commission",om["commission"]),
            ("Op Charges",om["op_charges"]),("Online Fee",om["online_fee"]),
            ("Total Payout",om["payout"])]]
        rows(ws_o,row,[[l,f"{v:,.2f}",f"{safe_div(v,om['delivered']):,.2f}",
            f"{safe_div(v,om['gmv'],pct=True):.1f}%"] for l,v in fin],{2,3,4})
        row=sec(ws_o,row+len(fin)+2,"  HOURLY ORDER PATTERN",color)
        hdr(ws_o,row,["Hour","Orders","Hour","Orders","Hour","Orders"],color,[10,10,10,10,10,10])
        row+=1
        hourly=sub_df.groupby("_hour")["Order ID"].count().reset_index().values.tolist()
        hr_rows=[]
        for i in range(0,len(hourly),3):
            trio=hourly[i:i+3]
            rv=[]
            for h,c_v in trio: rv+=[f"{int(h):02d}:00",int(c_v)]
            while len(rv)<6: rv+=["",""]
            hr_rows.append(rv)
        rows(ws_o,row,hr_rows,{2,4,6})

    # Raw Data
    ws_r=wb.create_sheet("Raw Data"); ws_r.sheet_view.showGridLines=False
    rc=[c for c in df.columns if not c.startswith("_")]
    rdf=df[rc]
    for ci,col in enumerate(rdf.columns,1):
        c=ws_r.cell(row=1,column=ci,value=col)
        c.font=fn(bold=True,size=7,color=C_WH); c.fill=fl(C_DARK); c.alignment=ct()
        ws_r.column_dimensions[get_column_letter(ci)].width=max(len(str(col))+2,10)
    for ri,rd in enumerate(rdf.itertuples(index=False),2):
        bg=C_LG if ri%2==0 else C_WH
        for ci,val in enumerate(rd,1):
            c=ws_r.cell(row=ri,column=ci,value=str(val) if pd.notna(val) else "")
            c.font=fn(size=7); c.fill=fl(bg)

    buf=BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ── PDF Builder ───────────────────────────────────────────────────────────────
def build_pdf(df):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle,PageBreak
    from reportlab.lib.enums import TA_CENTER,TA_LEFT

    buf=BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=1.5*cm,leftMargin=1.5*cm,
                          topMargin=1.5*cm,bottomMargin=1.5*cm)
    CD=colors.HexColor("#0F1117"); COR=colors.HexColor("#F97316")
    CGR=colors.HexColor("#10B981"); CRE=colors.HexColor("#EF4444")
    CBL=colors.HexColor("#3B82F6"); CWH=colors.white
    CLG=colors.HexColor("#F8FAFC")
    OCLR={"Liwan":COR,"Dubai Investment Park":CBL,"Oud Metha":colors.HexColor("#8B5CF6"),
          "Naif":CRE,"Al Muteena":CGR,"Al Hamriya":colors.HexColor("#F59E0B")}

    ts=ParagraphStyle("ts",fontSize=16,textColor=CWH,alignment=TA_CENTER,fontName="Helvetica-Bold",spaceAfter=4)
    sm=ParagraphStyle("sm",fontSize=7.5,textColor=colors.grey,fontName="Helvetica",alignment=TA_CENTER)
    W=17.5*cm

    def ban(txt,color=CD):
        d=[[Paragraph(txt,ts)]]; t=Table(d,colWidths=[W])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),color),
            ("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12)])); return t

    def sec_b(txt,color=COR):
        d=[[Paragraph(txt,ParagraphStyle("sb",fontSize=9,textColor=CWH,fontName="Helvetica-Bold"))]]
        t=Table(d,colWidths=[W])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),color),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),8)])); return t

    def tbl(data,cw,hc=COR):
        t=Table(data,colWidths=cw)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),hc),("TEXTCOLOR",(0,0),(-1,0),CWH),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.5),
            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[CWH,CLG]),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("ALIGN",(0,1),(0,-1),"LEFT"),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)])); return t

    story=[]; outlets_all=sorted(df["_area"].unique())
    delivered=df[df["Order status"]=="Delivered"]; cancelled=df[df["Order status"]=="Cancelled"]
    date_min=df["Order received at"].min(); date_max=df["Order received at"].max()
    outlet_data=[(a,outlet_metrics(df[df["_area"]==a])) for a in outlets_all if len(df[df["_area"]==a])>0]
    outlet_data.sort(key=lambda x:x[1]["gmv"],reverse=True)

    story.append(ban("AL MADINA HYPERMARKET",CD)); story.append(Spacer(1,.15*cm))
    story.append(ban("GROUP PERFORMANCE REPORT",COR)); story.append(Spacer(1,.15*cm))
    ds="" 
    if pd.notna(date_min) and pd.notna(date_max): ds=f"{date_min.strftime('%d %b %Y')} – {date_max.strftime('%d %b %Y')}"
    story.append(Paragraph(f"Period: {ds}   |   {len(outlet_data)} Outlets   |   Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",sm))
    story.append(Spacer(1,.4*cm))

    m=outlet_metrics(df); story.append(sec_b("GROUP KEY METRICS",CD)); story.append(Spacer(1,.1*cm))
    kd=[["Total Orders","Delivered","Cancelled","Cancel Rate","Gross Revenue","Total Payout"],
        [str(m["total"]),str(m["delivered"]),str(m["cancelled"]),f"{m['can_rate']:.1f}%",
         f"AED {m['gmv']:,.0f}",f"AED {m['payout']:,.0f}"],
        ["Avg Order","Commission","Avg Delivery","Pro Orders","Online Pay %","Complaints"],
        [f"AED {m['avg_order']:,.1f}",f"AED {m['commission']:,.0f}",
         f"{m['del_time']} min" if m['del_time'] else "N/A",
         f"{m['pro_pct']:.0f}%",f"{m['online_pct']:.0f}%",str(m["complaints"])]]
    kt=Table(kd,colWidths=[2.92*cm]*6)
    kt.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),CD),("BACKGROUND",(0,2),(-1,2),CD),
        ("TEXTCOLOR",(0,0),(-1,0),CWH),("TEXTCOLOR",(0,2),(-1,2),CWH),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTNAME",(0,2),(-1,2),"Helvetica-Bold"),
        ("FONTNAME",(0,1),(-1,1),"Helvetica"),("FONTNAME",(0,3),(-1,3),"Helvetica"),
        ("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story.append(kt); story.append(Spacer(1,.3*cm))

    story.append(sec_b("OUTLET LEAGUE TABLE  (by GMV)",COR)); story.append(Spacer(1,.1*cm))
    lt=[["#","Outlet","Orders","Delivered","Cancel%","GMV (AED)","Payout","Avg Order","Del Time"]]
    for rank,(area,om) in enumerate(outlet_data,1):
        lt.append([str(rank),area,str(om["total"]),str(om["delivered"]),
            f"{om['can_rate']:.1f}%",f"{om['gmv']:,.0f}",f"{om['payout']:,.0f}",
            f"{om['avg_order']:,.1f}",f"{om['del_time']:.0f}m" if om['del_time'] else "N/A"])
    ltt=tbl(lt,[.6*cm,3.5*cm,1.5*cm,1.8*cm,1.5*cm,2.2*cm,2.2*cm,1.8*cm,1.5*cm])
    for i,(area,_) in enumerate(outlet_data):
        ltt.setStyle(TableStyle([("BACKGROUND",(1,i+1),(1,i+1),OCLR.get(area,COR)),
                                  ("TEXTCOLOR",(1,i+1),(1,i+1),CWH)]))
    story.append(ltt); story.append(Spacer(1,.3*cm))

    story.append(sec_b("DAILY GROUP TREND",CBL)); story.append(Spacer(1,.1*cm))
    dg=df.groupby("_date").agg(total=("Order ID","count"),
        delivered=("Order status",lambda x:(x=="Delivered").sum()),
        cancelled=("Order status",lambda x:(x=="Cancelled").sum()),
        gmv=("Subtotal","sum")).reset_index()
    dg["can_pct"]=dg.apply(lambda r:safe_div(r["cancelled"],r["total"],pct=True),axis=1)
    dt=[["Date","Total","Delivered","Cancelled","Cancel %","GMV (AED)"]]+\
       [[str(r["_date"]),int(r["total"]),int(r["delivered"]),int(r["cancelled"]),
         f"{r['can_pct']:.1f}%",f"{r['gmv']:,.0f}"] for _,r in dg.iterrows()]
    story.append(tbl(dt,[3.5*cm,2.2*cm,2.2*cm,2.2*cm,2.5*cm,4.9*cm],CBL))
    story.append(PageBreak())

    for area,om in outlet_data:
        color=OCLR.get(area,COR)
        story.append(ban(f"OUTLET REPORT  ·  {area.upper()}",color)); story.append(Spacer(1,.2*cm))
        sub_df=df[df["_area"]==area]; sub_c=sub_df[sub_df["Order status"]=="Cancelled"]
        mk=[["Orders","Delivered","Cancelled","Cancel%","GMV","Avg Order","Del Time"],
            [str(om["total"]),str(om["delivered"]),str(om["cancelled"]),
             f"{om['can_rate']:.1f}%",f"AED {om['gmv']:,.0f}",
             f"AED {om['avg_order']:,.1f}",f"{om['del_time']:.0f}m" if om['del_time'] else "N/A"]]
        mkt=Table(mk,colWidths=[2.5*cm]*7)
        mkt.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),color),("TEXTCOLOR",(0,0),(-1,0),CWH),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTNAME",(0,1),(-1,1),"Helvetica"),
            ("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E2E8F0")),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
        story.append(mkt); story.append(Spacer(1,.2*cm))
        story.append(sec_b("Daily Performance",color)); story.append(Spacer(1,.08*cm))
        ds2=sub_df.groupby("_date").agg(total=("Order ID","count"),
            delivered=("Order status",lambda x:(x=="Delivered").sum()),
            cancelled=("Order status",lambda x:(x=="Cancelled").sum()),
            gmv=("Subtotal","sum")).reset_index()
        ds2["can_pct"]=ds2.apply(lambda r:safe_div(r["cancelled"],r["total"],pct=True),axis=1)
        dt2=[["Date","Total","Del","Can","Can%","GMV (AED)"]]+\
            [[str(r["_date"]),int(r["total"]),int(r["delivered"]),int(r["cancelled"]),
              f"{r['can_pct']:.1f}%",f"{r['gmv']:,.0f}"] for _,r in ds2.iterrows()]
        story.append(tbl(dt2,[3.2*cm,1.8*cm,1.8*cm,1.8*cm,2*cm,7.9*cm],color)); story.append(Spacer(1,.2*cm))
        if len(sub_c)>0:
            story.append(sec_b("Cancellations",CRE)); story.append(Spacer(1,.08*cm))
            cg=sub_c.groupby(["Cancellation owner","Cancellation reason"]).agg(
                Count=("Order ID","count"),Lost=("Subtotal","sum")).reset_index().sort_values("Count",ascending=False)
            ch=[["Owner","Reason","Count","Lost (AED)"]]+\
               [[str(r["Cancellation owner"]).strip(),str(r["Cancellation reason"])[:35],
                 int(r["Count"]),f"{r['Lost']:,.0f}"] for _,r in cg.iterrows()]
            story.append(tbl(ch,[2.5*cm,9*cm,1.8*cm,4.2*cm],CRE))
        story.append(Spacer(1,.2*cm))
        story.append(sec_b("Financial Breakdown",color)); story.append(Spacer(1,.08*cm))
        sub_d=sub_df[sub_df["Order status"]=="Delivered"]
        fh=[["Item","Total (AED)","Per Order","% of GMV"]]
        fi=[("GMV",om["gmv"]),("Commission",om["commission"]),
            ("Op Charges",om["op_charges"]),("Online Fee",om["online_fee"]),
            ("Payout",om["payout"])]
        fr=fh+[[l,f"{v:,.2f}",f"{safe_div(v,om['delivered']):,.2f}",
                f"{safe_div(v,om['gmv'],pct=True):.1f}%"] for l,v in fi]
        story.append(tbl(fr,[5*cm,4*cm,3.5*cm,5*cm],color))
        story.append(PageBreak())

    doc.build(story); buf.seek(0); return buf.read()

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    with st.sidebar:
        st.markdown("""
        <div style='text-align:center;padding:18px 0 10px;'>
            <div style='font-size:26px;margin-bottom:5px;'>📊</div>
            <div style='font-size:14px;font-weight:600;color:#f1f5f9;'>Al Madina Dashboard</div>
            <div style='font-size:11px;color:#475569;margin-top:2px;'>Performance Intelligence</div>
        </div>""", unsafe_allow_html=True)
        st.divider()
        uploaded = st.file_uploader("Upload Order Export (.xlsx)", type=["xlsx","xls"],
                                     help="Upload Talabat order details export")
        st.divider()

    if uploaded is None:
        st.markdown("""
        <div class="top-bar">
            <div class="top-bar-brand">
                <div class="logo-circle">AM</div>
                <div><h1>Al Madina Performance Dashboard</h1><span>Multi-Outlet Intelligence Platform</span></div>
            </div>
        </div>""", unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        for col,icon,title,desc,color in [
            (c1,"🏪","Multi-Outlet Analysis","Compare all 6 branches — league tables, ranking, benchmarks.","blue"),
            (c2,"📈","Full Analytics Suite","Sales, cancellations, delivery timing, operations, hourly patterns.","orange"),
            (c3,"📥","Management Reports","Branded Excel (per-outlet sheets) + PDF for sharing.","green"),
        ]:
            with col:
                st.markdown(f"""
                <div class="kpi-card {color}" style="padding:22px;">
                    <div style="font-size:26px;margin-bottom:8px;">{icon}</div>
                    <div style="font-size:14px;font-weight:600;color:#0f172a;margin-bottom:6px;">{title}</div>
                    <div style="font-size:12px;color:#64748b;">{desc}</div>
                </div>""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        alert("Upload your Talabat order details export (.xlsx) in the sidebar to launch the dashboard.", "blue", "👈")
        return

    with st.spinner("Loading & processing data..."):
        df = load_data(uploaded.read())
    if "Order status" not in df.columns:
        alert("Could not detect required columns. Please check your file format.", "red", "❌"); return

    outlets_all = sorted(df["_area"].unique())

    with st.sidebar:
        st.markdown("<div style='font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px;'>Date Range</div>", unsafe_allow_html=True)
        date_range = st.date_input("", value=(df["_date"].min(), df["_date"].max()),
                                   min_value=df["_date"].min(), max_value=df["_date"].max(),
                                   label_visibility="collapsed")
        st.markdown("<div style='font-size:10px;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em;margin:10px 0 6px;'>Outlets</div>", unsafe_allow_html=True)
        selected = st.multiselect("", options=outlets_all, default=outlets_all,
                                  label_visibility="collapsed")
        st.divider()
        st.markdown("<div style='font-size:10px;color:#475569;text-align:center;padding:6px;'>Al Madina Hypermarket<br>Performance Dashboard v2</div>", unsafe_allow_html=True)

    if len(date_range)==2:
        df = df[(df["_date"]>=date_range[0]) & (df["_date"]<=date_range[1])]
    if selected:
        df = df[df["_area"].isin(selected)]

    delivered = df[df["Order status"]=="Delivered"]
    cancelled = df[df["Order status"]=="Cancelled"]
    date_min = df["Order received at"].min()
    date_max = df["Order received at"].max()

    st.markdown(f"""
    <div class="top-bar">
        <div class="top-bar-brand">
            <div class="logo-circle">AM</div>
            <div>
                <h1>Al Madina Performance Dashboard</h1>
                <span>{len(selected)} outlet{"s" if len(selected)!=1 else ""} &nbsp;·&nbsp;
                {date_min.strftime('%d %b') if pd.notna(date_min) else '?'} – {date_max.strftime('%d %b %Y') if pd.notna(date_max) else '?'}
                </span>
            </div>
        </div>
        <div class="top-bar-meta">Generated {datetime.now().strftime('%d %b %Y, %H:%M')}</div>
    </div>""", unsafe_allow_html=True)

    tab1,tab2,tab3,tab4,tab5,tab6 = st.tabs([
        "🏢 Group Overview","🏪 Outlet Analysis",
        "💰 Sales & Revenue","❌ Cancellations","⚙️ Operations","📥 Download"])

    # ── Tab 1: Group Overview ──────────────────────────────────────────────
    with tab1:
        m = outlet_metrics(df)
        st.markdown('<div class="section-label">Group KPIs</div>', unsafe_allow_html=True)
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        with c1: kpi("Total Orders",m["total"],"","orange")
        with c2: kpi("Delivered",m["delivered"],f"{m['del_rate']:.1f}% rate","green")
        with c3: kpi("Cancelled",m["cancelled"],f"{m['can_rate']:.1f}% rate","red")
        with c4: kpi("Gross Revenue",f"AED {m['gmv']:,.0f}","","blue")
        with c5: kpi("Total Payout",f"AED {m['payout']:,.0f}","","purple")
        with c6: kpi("Avg Order",f"AED {m['avg_order']:,.1f}","","amber")
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        with c1: kpi("Commission",f"AED {m['commission']:,.0f}",f"{safe_div(m['commission'],m['gmv'],pct=True):.1f}% of GMV","orange")
        with c2: kpi("Op. Charges",f"AED {m['op_charges']:,.0f}","","blue")
        with c3: kpi("Avg Delivery",f"{m['del_time']} min" if m['del_time'] else "N/A","","green")
        with c4: kpi("Pro Orders",f"{m['pro_pct']:.0f}%","","purple")
        with c5: kpi("Online Pay",f"{m['online_pct']:.0f}%","","amber")
        with c6: kpi("Complaints",str(m["complaints"]),f"{m['complaint_rate']:.1f}%","red")

        st.markdown('<div class="section-label">Outlet League Table — Ranked by GMV</div>', unsafe_allow_html=True)
        outlet_data=[(a,outlet_metrics(df[df["_area"]==a])) for a in selected if len(df[df["_area"]==a])>0]
        outlet_data.sort(key=lambda x:x[1]["gmv"],reverse=True)
        max_gmv = max((o[1]["gmv"] for o in outlet_data),default=1)

        for rank,(area,om) in enumerate(outlet_data,1):
            color=OUTLET_COLORS.get(area,"#64748b")
            can_cls,can_lbl=can_badge(om["can_rate"])
            del_cls,del_lbl=del_badge(om["del_time"])
            c1,c2,c3=st.columns([2,3,2])
            with c1:
                store_id=str(df[df["_area"]==area]["Store ID"].iloc[0]) if len(df[df["_area"]==area])>0 else "N/A"
                st.markdown(f"""
                <div class="outlet-card">
                    <div class="outlet-header">
                        <div>
                            <div class="outlet-name">
                                <span style="font-size:20px;font-weight:700;color:{color};margin-right:6px;">#{rank}</span>{area}
                            </div>
                            <div class="outlet-area">Store {store_id}</div>
                        </div>
                        <div>
                            <span class="outlet-badge {can_cls}" style="display:block;margin-bottom:3px;">{can_lbl} cancel</span>
                            <span class="outlet-badge {del_cls}">{del_lbl} delivery</span>
                        </div>
                    </div>
                    <div style="font-size:22px;font-weight:600;color:#0f172a;">AED {om['gmv']:,.0f}</div>
                    <div style="font-size:12px;color:#94a3b8;margin-top:2px;">{om['total']} orders &nbsp;·&nbsp; {om['delivered']} delivered</div>
                </div>""", unsafe_allow_html=True)
            with c2:
                st.markdown('<div style="margin-top:8px;">', unsafe_allow_html=True)
                progress_bar("Orders delivered", om["delivered"], om["total"], color, f"/{om['total']}")
                progress_bar("GMV vs top outlet", round(om["gmv"]), round(max_gmv), color, " AED")
                progress_bar("Pro order share", round(om["pro_pct"]), 100, color, "%")
                progress_bar("Cancel rate", round(om["can_rate"],1), 60, "#ef4444" if om["can_rate"]>20 else "#f97316", "%")
                st.markdown('</div>', unsafe_allow_html=True)
            with c3:
                st.markdown(f"""
                <div style="margin-top:8px;">
                    <div class="compare-row">
                        <div class="compare-chip">Payout <span class="chip-val">AED {om['payout']:,.0f}</span></div>
                        <div class="compare-chip">Avg <span class="chip-val">AED {om['avg_order']:,.0f}</span></div>
                    </div>
                    <div class="compare-row">
                        <div class="compare-chip">Commission <span class="chip-val">AED {om['commission']:,.0f}</span></div>
                    </div>
                    <div class="compare-row">
                        <div class="compare-chip">Del time <span class="chip-val">{om['del_time']:.0f}m</span></div>
                        <div class="compare-chip">Cancel <span class="chip-val">{om['can_rate']:.1f}%</span></div>
                    </div>
                    <div class="compare-row">
                        <div class="compare-chip">Online <span class="chip-val">{om['online_pct']:.0f}%</span></div>
                        <div class="compare-chip">Pro <span class="chip-val">{om['pro_pct']:.0f}%</span></div>
                    </div>
                </div>""", unsafe_allow_html=True)

        st.markdown('<div class="section-label">Daily Trend</div>', unsafe_allow_html=True)
        dg=df.groupby("_date").agg(total=("Order ID","count"),
            delivered=("Order status",lambda x:(x=="Delivered").sum()),
            cancelled=("Order status",lambda x:(x=="Cancelled").sum()),
            gmv=("Subtotal","sum")).reset_index()
        dg.columns=["Date","Total","Delivered","Cancelled","GMV (AED)"]
        dg["Cancel %"]=dg.apply(lambda r:round(safe_div(r["Cancelled"],r["Total"],pct=True),1),axis=1)
        dg["GMV (AED)"]=dg["GMV (AED)"].round(0)
        st.dataframe(_style_df(dg, heat_col="GMV (AED)", heat_col2="Cancel %"),
                     use_container_width=True,hide_index=True)

        st.markdown('<div class="section-label">Key Insights</div>', unsafe_allow_html=True)
        if outlet_data:
            top=outlet_data[0]; low=outlet_data[-1]
            alert(f"<b>{top[0]}</b> leads the group with AED {top[1]['gmv']:,.0f} GMV — {safe_div(top[1]['gmv'],m['gmv'],pct=True):.0f}% of group total revenue.", "green", "🏆")
            high_can=max(outlet_data,key=lambda x:x[1]["can_rate"])
            if high_can[1]["can_rate"]>15:
                alert(f"<b>{high_can[0]}</b> has the highest cancellation rate at <b>{high_can[1]['can_rate']:.1f}%</b> — action needed. Lost AED {high_can[1]['lost_gmv']:,.0f}.", "red", "🚨")
            slow=[f"{a} ({o['del_time']:.0f}m)" for a,o in outlet_data if o["del_time"] and o["del_time"]>42]
            if slow:
                alert(f"Outlets with slow delivery (>42 min): <b>{', '.join(slow)}</b>. Review operations.", "amber", "⏱")
            alert(f"<b>{low[0]}</b> has the lowest GMV at AED {low[1]['gmv']:,.0f} with {low[1]['total']} orders. Growth opportunity.", "blue", "📌")
            if m["can_rate"]>10:
                alert(f"Group cancellation rate is <b>{m['can_rate']:.1f}%</b>. Total revenue lost: <b>AED {m['lost_gmv']:,.0f}</b>. Primary cause: Item not available.", "amber", "💸")

    # ── Tab 2: Outlet Analysis ─────────────────────────────────────────────
    with tab2:
        sel_area = st.selectbox("Select outlet", options=[a for a in outlets_all if a in selected])
        sub_df=df[df["_area"]==sel_area]; sub_d=sub_df[sub_df["Order status"]=="Delivered"]
        sub_c=sub_df[sub_df["Order status"]=="Cancelled"]; om=outlet_metrics(sub_df)
        color=OUTLET_COLORS.get(sel_area,"#64748b")
        can_cls,can_lbl=can_badge(om["can_rate"]); del_cls,del_lbl=del_badge(om["del_time"])
        fname=sub_df["Restaurant name"].iloc[0] if len(sub_df)>0 else sel_area
        st.markdown(f"""
        <div style="background:white;border:1px solid #e2e8f0;border-radius:14px;padding:18px 22px;
                    margin-bottom:16px;border-left:4px solid {color};">
            <div style="font-size:17px;font-weight:600;color:#0f172a;">{sel_area}</div>
            <div style="font-size:11px;color:#94a3b8;margin-top:2px;">{fname}</div>
            <div style="display:flex;gap:8px;margin-top:8px;flex-wrap:wrap;">
                <span class="outlet-badge {can_cls}">Cancel: {can_lbl} ({om['can_rate']:.1f}%)</span>
                <span class="outlet-badge {del_cls}">Delivery: {del_lbl} ({f"{om['del_time']:.0f} min" if om['del_time'] else 'N/A'})</span>
            </div>
        </div>""", unsafe_allow_html=True)
        c1,c2,c3,c4,c5,c6=st.columns(6)
        with c1: kpi("Orders",om["total"],"","orange")
        with c2: kpi("Delivered",om["delivered"],f"{om['del_rate']:.1f}%","green")
        with c3: kpi("Cancelled",om["cancelled"],f"{om['can_rate']:.1f}%","red")
        with c4: kpi("GMV",f"AED {om['gmv']:,.0f}","","blue")
        with c5: kpi("Payout",f"AED {om['payout']:,.0f}","","purple")
        with c6: kpi("Avg Order",f"AED {om['avg_order']:,.1f}","","amber")
        c1,c2,c3,c4=st.columns(4)
        with c1: kpi("Avg Delivery",f"{om['del_time']:.0f} min" if om['del_time'] else "N/A","","orange")
        with c2: kpi("Avg Prep",f"{om['prep_time']:.0f} min" if om['prep_time'] else "N/A","","blue")
        with c3: kpi("Last Mile",f"{om['last_mile']:.0f} min" if om['last_mile'] else "N/A","","green")
        with c4: kpi("Complaints",str(om["complaints"]),f"{om['complaint_rate']:.1f}%","red")

        c1,c2=st.columns(2)
        with c1:
            st.markdown('<div class="section-label">Daily Performance</div>',unsafe_allow_html=True)
            ds=sub_df.groupby("_date").agg(total=("Order ID","count"),
                delivered=("Order status",lambda x:(x=="Delivered").sum()),
                cancelled=("Order status",lambda x:(x=="Cancelled").sum()),
                gmv=("Subtotal","sum"),payout=("Payout Amount","sum")).reset_index()
            ds.columns=["Date","Total","Delivered","Cancelled","GMV (AED)","Payout (AED)"]
            ds["Cancel %"]=ds.apply(lambda r:round(safe_div(r["Cancelled"],r["Total"],pct=True),1),axis=1)
            ds["GMV (AED)"]=ds["GMV (AED)"].round(0); ds["Payout (AED)"]=ds["Payout (AED)"].round(0)
            st.dataframe(_style_df(ds, heat_col="GMV (AED)"),
                         use_container_width=True,hide_index=True)
        with c2:
            st.markdown('<div class="section-label">Financial Breakdown</div>',unsafe_allow_html=True)
            fin={"GMV":om["gmv"],"Commission":om["commission"],"Op Charges":om["op_charges"],
                 "Online Fee":om["online_fee"],"Total Payout":om["payout"]}
            fin_df=pd.DataFrame([{"Item":k,"Total":round(v,2),
                "Per Order":round(safe_div(v,om["delivered"]),2),
                "% of GMV":round(safe_div(v,om["gmv"],pct=True),1)} for k,v in fin.items()])
            st.dataframe(fin_df,use_container_width=True,hide_index=True)

        if len(sub_c)>0:
            st.markdown('<div class="section-label">Cancellation Detail</div>',unsafe_allow_html=True)
            cd=sub_c.groupby(["Cancellation owner","Cancellation reason"]).agg(
                Count=("Order ID","count"),Lost_GMV=("Subtotal","sum")).reset_index().sort_values("Count",ascending=False)
            cd.columns=["Owner","Reason","Count","Lost GMV (AED)"]
            cd["Lost GMV (AED)"]=cd["Lost GMV (AED)"].round(2)
            st.dataframe(cd,use_container_width=True,hide_index=True)
        else:
            alert(f"No cancellations for {sel_area} in this period.","green","✅")

        c1,c2=st.columns(2)
        with c1:
            st.markdown('<div class="section-label">Hourly Order Pattern</div>',unsafe_allow_html=True)
            h=sub_df.groupby("_hour")["Order ID"].count().reset_index()
            h.columns=["Hour","Orders"]; h["Hour"]=h["Hour"].apply(lambda x:f"{x:02d}:00")
            st.dataframe(_style_df(h, heat_col="Orders"),
                         use_container_width=True,hide_index=True)
        with c2:
            st.markdown('<div class="section-label">Day of Week Pattern</div>',unsafe_allow_html=True)
            dow_order=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            dw=sub_df.groupby("_dow")["Order ID"].count().reindex(dow_order).reset_index()
            dw.columns=["Day","Orders"]; dw["Orders"]=dw["Orders"].fillna(0).astype(int)
            st.dataframe(_style_df(dw, heat_col="Orders"),
                         use_container_width=True,hide_index=True)

    # ── Tab 3: Sales ───────────────────────────────────────────────────────
    with tab3:
        st.markdown('<div class="section-label">Revenue by Outlet</div>',unsafe_allow_html=True)
        rev=[]
        for area in selected:
            sd=delivered[delivered["_area"]==area]
            s=df[df["_area"]==area]
            if len(s)==0: continue
            rev.append({"Outlet":area,"Orders":len(s),"Delivered":len(sd),
                "GMV (AED)":round(sd["Subtotal"].sum(),0),
                "Payout (AED)":round(sd["Payout Amount"].sum(),0),
                "Commission (AED)":round(sd["Commission"].sum(),0),
                "Avg Order":round(safe_div(sd["Subtotal"].sum(),len(sd)),1),
                "Pro %":round(safe_div(len(sd[sd["Is Pro Order"]=="Y"]),len(sd),pct=True),1),
                "Online %":round(safe_div(len(sd[sd["Payment type"]=="Online"]),len(sd),pct=True),1)})
        rev_df=pd.DataFrame(rev).sort_values("GMV (AED)",ascending=False)
        st.dataframe(_style_df(rev_df, heat_col="GMV (AED)", heat_col2="Avg Order"),
                     use_container_width=True,hide_index=True)

        c1,c2=st.columns(2)
        with c1:
            st.markdown('<div class="section-label">GMV by Outlet by Day</div>',unsafe_allow_html=True)
            gp=delivered.groupby(["_date","_area"])["Subtotal"].sum().unstack(fill_value=0).round(0)
            st.dataframe(_style_df_full(gp),use_container_width=True)
        with c2:
            st.markdown('<div class="section-label">Day-of-Week Pattern</div>',unsafe_allow_html=True)
            dow_order=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            dw=df.groupby("_dow").agg(Orders=("Order ID","count"),GMV=("Subtotal","sum")).reindex(dow_order).reset_index()
            dw.columns=["Day","Total Orders","GMV (AED)"]; dw["GMV (AED)"]=dw["GMV (AED)"].round(0)
            st.dataframe(_style_df(dw, heat_col="Total Orders"),
                         use_container_width=True,hide_index=True)

        st.markdown('<div class="section-label">Payment Mix by Outlet</div>',unsafe_allow_html=True)
        pay=[]
        for area in selected:
            sd=delivered[delivered["_area"]==area]
            if len(sd)==0: continue
            on=sd[sd["Payment type"]=="Online"]; ca=sd[sd["Payment type"]=="Cash"]
            pay.append({"Outlet":area,"Online Orders":len(on),"Cash Orders":len(ca),
                "Online GMV":round(on["Subtotal"].sum(),0),"Cash GMV":round(ca["Subtotal"].sum(),0),
                "Online %":round(safe_div(len(on),len(sd),pct=True),1)})
        st.dataframe(pd.DataFrame(pay),use_container_width=True,hide_index=True)

    # ── Tab 4: Cancellations ──────────────────────────────────────────────
    with tab4:
        st.markdown('<div class="section-label">Cancellation Scorecard</div>',unsafe_allow_html=True)
        can=[]
        for area in selected:
            s=df[df["_area"]==area]; sc=s[s["Order status"]=="Cancelled"]
            if len(s)==0: continue
            ow=sc["Cancellation owner"].value_counts()
            can.append({"Outlet":area,"Total":len(s),"Cancelled":len(sc),
                "Cancel Rate %":round(safe_div(len(sc),len(s),pct=True),1),
                "Vendor":int(ow.get("Vendor",0)),"Customer":int(ow.get("Customer",0)),
                "Rider":int(ow.get("Rider",0)),
                "Item N/A":int(sc["Cancellation reason"].str.contains("Item not available",na=False).sum()),
                "Fraudulent":int(sc["Cancellation reason"].str.contains("Fraudulent",na=False).sum()),
                "Lost GMV (AED)":round(sc["Subtotal"].sum(),0)})
        can_df=pd.DataFrame(can).sort_values("Cancel Rate %",ascending=False)
        st.dataframe(_style_df(can_df, heat_col="Lost GMV (AED)", heat_col2="Cancel Rate %"),
                     use_container_width=True,hide_index=True)
        st.markdown('<div class="section-label">Outlet Alerts</div>',unsafe_allow_html=True)
        for _,r in can_df.iterrows():
            if r["Cancel Rate %"]>40:
                alert(f"<b>{r['Outlet']}</b>: CRITICAL — {r['Cancel Rate %']:.1f}% cancellation rate. {r['Cancelled']}/{r['Total']} orders cancelled. Lost AED {r['Lost GMV (AED)']:,.0f}. Vendor-caused: {r['Vendor']}","red","🚨")
            elif r["Cancel Rate %"]>10:
                primary="item availability" if r["Item N/A"]>r["Fraudulent"] else "fraudulent orders"
                alert(f"<b>{r['Outlet']}</b>: Elevated cancel rate {r['Cancel Rate %']:.1f}%. Main cause: {primary}. Review with outlet manager.","amber","⚠️")
            else:
                alert(f"<b>{r['Outlet']}</b>: Good cancel rate {r['Cancel Rate %']:.1f}%.","green","✅")
        if len(cancelled)>0:
            st.markdown('<div class="section-label">All Cancellations Detail</div>',unsafe_allow_html=True)
            all_c=cancelled.groupby(["_area","Cancellation owner","Cancellation reason"]).agg(
                Count=("Order ID","count"),Lost=("Subtotal","sum")).reset_index().sort_values("Count",ascending=False)
            all_c.columns=["Outlet","Owner","Reason","Count","Lost GMV (AED)"]
            all_c["Lost GMV (AED)"]=all_c["Lost GMV (AED)"].round(2)
            st.dataframe(all_c,use_container_width=True,hide_index=True)

    # ── Tab 5: Operations ─────────────────────────────────────────────────
    with tab5:
        st.markdown('<div class="section-label">Delivery Timing Comparison</div>',unsafe_allow_html=True)
        tm=[]
        for area in selected:
            sd=delivered[delivered["_area"]==area]
            if len(sd)==0: continue
            tot=(sd["Delivered at"]-sd["Order received at"]).dt.total_seconds()/60; tot=tot[tot>0]
            prep=(sd["Ready to pick up at"]-sd["Accepted at"]).dt.total_seconds()/60; prep=prep[prep>0]
            lm=(sd["Delivered at"]-sd["In delivery at"]).dt.total_seconds()/60; lm=lm[lm>0]
            tm.append({"Outlet":area,"Avg Total (min)":round(tot.mean(),1) if len(tot)>0 else None,
                "Avg Prep (min)":round(prep.mean(),1) if len(prep)>0 else None,
                "Avg Last Mile (min)":round(lm.mean(),1) if len(lm)>0 else None,
                "Max (min)":round(tot.max(),1) if len(tot)>0 else None,
                "Min (min)":round(tot.min(),1) if len(tot)>0 else None,
                "Deliveries":len(tot)})
        tm_df=pd.DataFrame(tm).sort_values("Avg Total (min)")
        st.dataframe(_style_df(tm_df, heat_col="Avg Total (min)", reverse=True),
                     use_container_width=True,hide_index=True)
        st.markdown('<div class="section-label">Timing Alerts</div>',unsafe_allow_html=True)
        for _,r in tm_df.iterrows():
            v=r["Avg Total (min)"]
            if v is None: continue
            if v<=33: alert(f"<b>{r['Outlet']}</b>: Excellent delivery — {v} min avg (prep {r['Avg Prep (min)']}m + last mile {r['Avg Last Mile (min)']}m).","green","🚀")
            elif v<=42: alert(f"<b>{r['Outlet']}</b>: Acceptable delivery — {v} min avg. Benchmark is 35 min.","blue","🕐")
            else: alert(f"<b>{r['Outlet']}</b>: Slow delivery — {v} min avg. Prep: {r['Avg Prep (min)']}m, Last mile: {r['Avg Last Mile (min)']}m. Investigate bottleneck.","red","⚠️")
        c1,c2=st.columns(2)
        with c1:
            st.markdown('<div class="section-label">Operational Costs</div>',unsafe_allow_html=True)
            op=[]
            for area in selected:
                sd=delivered[delivered["_area"]==area]
                if len(sd)==0: continue
                tc=sd["Commission"].sum()+sd["Operational Charges"].sum()+sd["Online Payment Fee"].sum()
                op.append({"Outlet":area,"Commission":round(sd["Commission"].sum(),0),
                    "Op Charges":round(sd["Operational Charges"].sum(),0),
                    "Online Fee":round(sd["Online Payment Fee"].sum(),0),
                    "Total Cost":round(tc,0),"Cost/Order":round(safe_div(tc,len(sd)),1)})
            st.dataframe(pd.DataFrame(op).sort_values("Total Cost",ascending=False)
                           ,
                         use_container_width=True,hide_index=True)
        with c2:
            st.markdown('<div class="section-label">Hourly Demand Heatmap</div>',unsafe_allow_html=True)
            hp=df.groupby(["_hour","_area"])["Order ID"].count().unstack(fill_value=0)
            hp.index=[f"{h:02d}:00" for h in hp.index]
            st.dataframe(_style_df_full(hp),use_container_width=True)
        st.markdown('<div class="section-label">Complaint Analysis</div>',unsafe_allow_html=True)
        comp=[]
        for area in selected:
            sd=delivered[delivered["_area"]==area]
            if len(sd)==0: continue
            c_comp=sd[sd["Has Complaint?"]=="Y"]
            comp.append({"Outlet":area,"Deliveries":len(sd),"Complaints":len(c_comp),
                "Rate %":round(safe_div(len(c_comp),len(sd),pct=True),2),
                "Reasons":", ".join(c_comp["Complaint Reason"].dropna().unique()) if len(c_comp)>0 else "—"})
        st.dataframe(pd.DataFrame(comp),use_container_width=True,hide_index=True)

    # ── Tab 6: Download ───────────────────────────────────────────────────
    with tab6:
        st.markdown('<div class="section-label">Export Reports</div>',unsafe_allow_html=True)
        c1,c2=st.columns(2)
        with c1:
            st.markdown("""
            <div class="kpi-card orange" style="padding:22px;margin-bottom:14px;">
                <div style="font-size:22px;margin-bottom:8px;">📊</div>
                <div style="font-size:15px;font-weight:600;color:#0f172a;margin-bottom:6px;">Excel Report</div>
                <div style="font-size:12px;color:#64748b;margin-bottom:10px;">Full workbook — one sheet per outlet plus group summary, comparison matrix, and raw data.</div>
                <ul style="font-size:11px;color:#64748b;margin-left:14px;line-height:1.8;">
                    <li>Group Summary + League Table</li>
                    <li>Outlet Comparison Matrix</li>
                    <li>Individual outlet sheets (daily, financial, cancellations, hourly)</li>
                    <li>Raw data sheet</li>
                </ul>
            </div>""", unsafe_allow_html=True)
            if st.button("📥 Generate & Download Excel", use_container_width=True, type="primary", key="gen_excel"):
                try:
                    with st.spinner("Building Excel report..."):
                        excel_bytes = build_excel(df)
                    st.download_button(
                        "💾 Click here to save Excel",
                        data=excel_bytes,
                        file_name=f"AlMadina_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Excel build failed: {e}. Check that openpyxl is in requirements.txt")
        with c2:
            st.markdown("""
            <div class="kpi-card blue" style="padding:22px;margin-bottom:14px;">
                <div style="font-size:22px;margin-bottom:8px;">📄</div>
                <div style="font-size:15px;font-weight:600;color:#0f172a;margin-bottom:6px;">PDF Report</div>
                <div style="font-size:12px;color:#64748b;margin-bottom:10px;">Print-ready management report for presentations and outlet manager distribution.</div>
                <ul style="font-size:11px;color:#64748b;margin-left:14px;line-height:1.8;">
                    <li>Cover page + Group KPIs</li>
                    <li>Outlet league table</li>
                    <li>Daily trend</li>
                    <li>Per-outlet pages</li>
                </ul>
            </div>""", unsafe_allow_html=True)
            if st.button("📄 Generate & Download PDF", use_container_width=True, key="gen_pdf"):
                try:
                    with st.spinner("Building PDF report..."):
                        pdf_bytes = build_pdf(df)
                    st.download_button(
                        "💾 Click here to save PDF",
                        data=pdf_bytes,
                        file_name=f"AlMadina_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"PDF build failed: {e}. Check that reportlab is in requirements.txt")
        st.divider()
        m=outlet_metrics(df)
        alert(f"Report covers <b>{len(selected)} outlets</b>, <b>{m['total']} orders</b> ({m['delivered']} delivered, {m['cancelled']} cancelled), <b>AED {m['gmv']:,.0f}</b> gross revenue — period <b>{date_min.strftime('%d %b') if pd.notna(date_min) else '?'} – {date_max.strftime('%d %b %Y') if pd.notna(date_max) else '?'}</b>.","blue","📋")

if __name__=="__main__":
    main()
